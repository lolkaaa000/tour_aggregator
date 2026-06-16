#!/usr/bin/env python3
"""
Генерация тестовых данных для турагрегатора.
Сценарий создаёт:
1. Схему БД и справочники
2. Нормализованные туры и отели
3. Источники операторов в CSV / JSON / XLSX
4. Импорт сырых данных из файлов в raw-слой
5. Пользовательские события и бронирования
"""

from __future__ import annotations

import csv
import json
import os
import random
import shutil
import sqlite3
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "download", "tour_aggregator.db")
SCHEMA_PATH = os.path.join(os.path.dirname(__file__), "..", "sql", "01_schema.sql")
SOURCE_DIR = os.path.join(os.path.dirname(__file__), "..", "download", "sources")

COUNTRIES = [
    ("Турция", "TR", "Средиземноморье", 1),
    ("Египет", "EG", "Африка", 1),
    ("Таиланд", "TH", "Юго-Восточная Азия", 1),
    ("ОАЭ", "AE", "Ближний Восток", 1),
    ("Вьетнам", "VN", "Юго-Восточная Азия", 1),
    ("Шри-Ланка", "LK", "Южная Азия", 0),
    ("Грузия", "GE", "Кавказ", 1),
    ("Кипр", "CY", "Средиземноморье", 1),
    ("Греция", "GR", "Европа", 1),
    ("Испания", "ES", "Европа", 1),
    ("Италия", "IT", "Европа", 0),
    ("Черногория", "ME", "Европа", 0),
    ("Доминикана", "DO", "Карибы", 1),
    ("Мексика", "MX", "Центральная Америка", 0),
    ("Мальдивы", "MV", "Индийский океан", 1),
    ("Тунис", "TN", "Африка", 0),
    ("Марокко", "MA", "Африка", 0),
    ("Абхазия", "AB", "Кавказ", 0),
    ("Россия", "RU", "Европа/Азия", 1),
    ("Куба", "CU", "Карибы", 0),
]

RESORTS_BY_COUNTRY = {
    "Турция": ["Анталья", "Аланья", "Кемер", "Сиде", "Белек", "Фетхие", "Мармарис", "Бодрум", "Стамбул"],
    "Египет": ["Хургада", "Шарм-эль-Шейх", "Марса-Алам", "Дахаб", "Таба"],
    "Таиланд": ["Паттайя", "Пхукет", "Самуи", "Краби", "Хуахин"],
    "ОАЭ": ["Дубай", "Шарджа", "Абу-Даби", "Рас-эль-Хайма", "Фуджейра"],
    "Вьетнам": ["Нячанг", "Фантхиет", "Дананг", "Фукуок", "Хойан"],
    "Шри-Ланка": ["Бентота", "Унаватуна", "Тангалле", "Тринкомали", "Калутара"],
    "Грузия": ["Батуми", "Кобулети", "Уреки", "Местиа", "Тбилиси"],
    "Кипр": ["Лимасол", "Ларнака", "Пафос", "Айя-Напа", "Протарас"],
    "Греция": ["Крит", "Родос", "Корфу", "Санторини", "Закинф"],
    "Испания": ["Коста-дель-Соль", "Тенерифе", "Майорка", "Барселона", "Ибица"],
    "Италия": ["Римини", "Сицилия", "Сардиния", "Неаполь", "Венеция"],
    "Черногория": ["Будва", "Котор", "Бечичи", "Тиват", "Святой Стефан"],
    "Доминикана": ["Пунта-Кана", "Пуэрто-Плата", "Ла-Романа", "Самана", "Баваро"],
    "Мексика": ["Канкун", "Ривьера-Майя", "Лос-Кабос", "Пуэрто-Вальярта"],
    "Мальдивы": ["Мале", "Ари-Атолл", "Баа-Атолл", "Лхавияни-Атолл"],
    "Тунис": ["Сусс", "Хаммамет", "Монастир", "Джерба", "Сфакс"],
    "Марокко": ["Агадир", "Марракеш", "Эссуэйра", "Танжер"],
    "Абхазия": ["Гагра", "Пицунда", "Сухум", "Новый Афон"],
    "Россия": ["Сочи", "Анапа", "Геленджик", "Ялта", "Калининград"],
    "Куба": ["Варадеро", "Гавана", "Тринидад", "Кайо-Коко"],
}

OPERATORS = [
    ("Библио-Глобус", "BG", 1, "https://bg-operator.ru"),
    ("Пегас Туристик", "Pegas", 1, "https://pegast.ru"),
    ("Coral Travel", "Coral", 1, "https://coral.ru"),
    ("Anex Tour", "Anex", 1, "https://anextour.ru"),
    ("FUN&SUN", "FUNSUN", 1, "https://funsun.ru"),
    ("TUI Россия", "TUI", 1, "https://tui.ru"),
    ("Спутник", "Sputnik", 1, "https://sputnik.ru"),
    ("МТС Travel", "MTS", 1, "https://mts-travel.ru"),
]

OPERATOR_RAW_NAMES = {
    "Библио-Глобус": ["Библио-Глобус", "Biblio Globus", "БИБЛИО ГЛОБУС", "biblio-globus", "Б.Глобус"],
    "Пегас Туристик": ["Пегас Туристик", "Pegas Touristik", "Пегас", "PEGAS", "Пегас-Туристик"],
    "Coral Travel": ["Coral Travel", "Корал Тревел", "CORAL", "Coral", "Корал Трэвел"],
    "Anex Tour": ["Anex Tour", "Анекс Тур", "ANEX", "Анекс", "Anex"],
    "FUN&SUN": ["FUN&SUN", "Фан энд Сан", "FUN AND SUN", "FunSun", "ФанСан"],
    "TUI Россия": ["TUI Россия", "TUI Russia", "ТЮИ", "TUI", "TUI Россия"],
    "Спутник": ["Спутник", "Sputnik", "Sputnik Tour", "СПУТНИК", "Sputnik-Тур"],
    "МТС Travel": ["МТС Travel", "МТС Тревел", "MTS Travel", "МТС", "MTS"],
}

COUNTRY_VARIANTS = {
    "Турция": ["Турция", "Турций", "Turkey", "Türkiye"],
    "Египет": ["Египет", "Египт", "Egypt"],
    "Таиланд": ["Таиланд", "Таеланд", "Thailand", "Тайланд"],
    "ОАЭ": ["ОАЭ", "ОАЕ", "UAE", "Объединенные Арабские Эмираты"],
    "Вьетнам": ["Вьетнам", "Vietnam"],
    "Шри-Ланка": ["Шри-Ланка", "Шри Ланка", "Sri Lanka"],
    "Грузия": ["Грузия", "Georgia"],
    "Кипр": ["Кипр", "Cyprus"],
    "Греция": ["Греция", "Greece"],
    "Испания": ["Испания", "Spain"],
    "Италия": ["Италия", "Italy"],
    "Черногория": ["Черногория", "Montenegro"],
    "Доминикана": ["Доминикана", "Dominican Republic", "Доминиканская Республика"],
    "Мексика": ["Мексика", "Mexico"],
    "Мальдивы": ["Мальдивы", "Maldives"],
    "Тунис": ["Тунис", "Tunisia"],
    "Марокко": ["Марокко", "Morocco"],
    "Абхазия": ["Абхазия"],
    "Россия": ["Россия", "Russia"],
    "Куба": ["Куба", "Cuba"],
}

RESORT_VARIANTS = {
    "Шарм-эль-Шейх": ["Шарм-эль-Шейх", "Шарм эль Шейх", "Sharm El Sheikh", "Шарм"],
    "Рас-эль-Хайма": ["Рас-эль-Хайма", "Рас эль Хайма", "Ras Al Khaimah"],
    "Коста-дель-Соль": ["Коста-дель-Соль", "Коста дель Соль", "Costa del Sol"],
    "Пунта-Кана": ["Пунта-Кана", "Пунта Кана", "Punta Cana"],
    "Ривьера-Майя": ["Ривьера-Майя", "Ривьера Майя", "Riviera Maya"],
    "Новый Афон": ["Новый Афон", "Новый  Афон"],
}

MEAL_TYPES = ["AI", "UAI", "FB", "HB", "BB", "OB", "RO"]
MEAL_VARIANTS = {
    "AI": ["AI", "All Inclusive", "all inclusive", "всё включено"],
    "UAI": ["UAI", "Ultra All Inclusive", "ультра всё включено"],
    "FB": ["FB", "Full Board", "полный пансион", "FB+"],
    "HB": ["HB", "Half Board", "полупансион", "питание"],
    "BB": ["BB", "Bed & Breakfast", "завтраки", "завтрак"],
    "OB": ["OB", "Room Only", "только кровать"],
    "RO": ["RO", "Without Meal", "без питания"],
}

BASE_PRICES = {
    "Мальдивы": 12500, "ОАЭ": 10000, "Доминикана": 9000, "Мексика": 8500,
    "Таиланд": 6000, "Куба": 6500, "Шри-Ланка": 5000, "Вьетнам": 4500,
    "Турция": 4000, "Египет": 3500, "Кипр": 5500, "Греция": 6500,
    "Испания": 7000, "Италия": 7500, "Черногория": 4000, "Грузия": 3000,
    "Тунис": 2750, "Марокко": 3250, "Абхазия": 2000, "Россия": 3500,
}

STAR_MULTIPLIERS = {2: 0.5, 3: 0.7, 4: 1.0, 5: 1.5}
MEAL_MULTIPLIERS = {"RO": 0.6, "OB": 0.7, "BB": 0.8, "HB": 0.9, "FB": 1.1, "AI": 1.2, "UAI": 1.4}
OPERATOR_DISCOUNTS = {"BG": 0.0, "Pegas": 0.03, "Coral": 0.05, "Anex": 0.02, "FUNSUN": 0.04, "TUI": 0.06, "Sputnik": 0.01, "MTS": 0.07}
ROOM_TYPES = ["Standard", "Superior", "Deluxe", "Suite", "Family Room", "Bungalow", "Villa", "Studio"]
BOOKING_STATUSES = ["pending", "confirmed", "cancelled", "completed"]
AGE_GROUPS = ["18-24", "25-34", "35-44", "45-54", "55-64", "65+"]
CITIES = ["Москва", "Санкт-Петербург", "Казань", "Новосибирск", "Екатеринбург", "Нижний Новгород", "Самара", "Ростов-на-Дону", "Уфа", "Воронеж", "Краснодар", "Челябинск", "Пермь", "Волгоград", "Омск"]
HOTEL_NAME_PARTS = {
    "prefix": ["Hotel", "Resort", "Grand", "Royal", "Imperial", "Palace", "Villa", "Beach", "Spa", "Boutique"],
    "name": ["Palace", "Paradise", "Sunset", "Oasis", "Azure", "Marina", "Garden", "Bay", "Crown", "Star", "Diamond", "Emerald", "Pearl", "Sapphire", "Golden", "Silver", "Crystal", "Coral", "Palm", "Lagoon", "Horizon", "Mirage", "Phoenix", "Serenity", "Harmony", "Lotus", "Orchid", "Dolphin", "Turtle", "Pelican"],
}


def log(message):
    print(message)


def random_date(start, end):
    return start + timedelta(days=random.randint(0, (end - start).days))


def generate_tour_price(country, stars, meal, nights, operator_short):
    base = BASE_PRICES.get(country, 3500)
    total = base * STAR_MULTIPLIERS.get(stars, 1.0) * MEAL_MULTIPLIERS.get(meal, 1.0) * nights * 2
    total *= 1 - OPERATOR_DISCOUNTS.get(operator_short, 0)
    total *= 1 + random.gauss(0, 0.08)
    return round(max(total, 25000), 0)


def safe_reset_path(path):
    abs_path = os.path.abspath(path)
    if os.path.isdir(abs_path):
        shutil.rmtree(abs_path)
    elif os.path.exists(abs_path):
        os.remove(abs_path)


def create_variant_hotel_name(name):
    variant = name
    replacements = [("Hotel", "Htl"), ("Resort", "Rsort"), ("Grand", "Grnd"), ("Beach", "Bch")]
    for old, new in replacements:
        if old in variant and random.random() < 0.5:
            variant = variant.replace(old, new)
    if random.random() < 0.2:
        variant = variant.upper()
    return variant


def choose_variant(mapping, canonical):
    variants = mapping.get(canonical, [canonical])
    return random.choice(variants)


def write_csv(path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path, rows):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, ensure_ascii=False, indent=2)


def write_xlsx(path, sheet_name, rows, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    wb.save(path)


def load_rows_from_source(file_path, source_type, entity):
    if source_type == "csv":
        with open(file_path, "r", encoding="utf-8-sig") as fh:
            return list(csv.DictReader(fh))
    if source_type == "json":
        with open(file_path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    if source_type == "xlsx":
        wb = load_workbook(file_path, read_only=True)
        ws = wb[entity]
        headers = [cell for cell in next(ws.iter_rows(values_only=True))]
        return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    raise ValueError(f"Неподдерживаемый формат: {source_type}")


def import_source_rows(conn, operator_id, operator_name, source_type, entity, file_path, batch_name):
    started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO source_import_runs (
            operator_id, source_name, source_type, file_path, status, started_at
        ) VALUES (?, ?, ?, ?, 'running', ?)
        """,
        (operator_id, f"{operator_name}_{entity}", source_type, file_path, started_at),
    )
    run_id = cur.lastrowid

    loaded = 0
    failed = 0
    rows = load_rows_from_source(file_path, source_type, entity)
    for row in rows:
        try:
            if entity == "hotels":
                mapped = map_hotel_row(row, source_type)
                cur.execute(
                    """
                    INSERT INTO hotels_raw (
                        operator_id, raw_name, raw_country, raw_resort, raw_stars, raw_meal,
                        source_name, source_type, load_batch
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        operator_id,
                        mapped["raw_name"],
                        mapped["raw_country"],
                        mapped["raw_resort"],
                        mapped["raw_stars"],
                        mapped["raw_meal"],
                        f"{operator_name}_{entity}",
                        source_type,
                        batch_name,
                    ),
                )
            else:
                mapped = map_tour_row(row, source_type)
                cur.execute(
                    """
                    INSERT INTO tours_raw (
                        operator_id, raw_hotel_name, raw_country, raw_resort, raw_price,
                        raw_departure, raw_duration, raw_meal, raw_adults, raw_children,
                        source_name, source_type, load_batch
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        operator_id,
                        mapped["raw_hotel_name"],
                        mapped["raw_country"],
                        mapped["raw_resort"],
                        mapped["raw_price"],
                        mapped["raw_departure"],
                        mapped["raw_duration"],
                        mapped["raw_meal"],
                        mapped["raw_adults"],
                        mapped["raw_children"],
                        f"{operator_name}_{entity}",
                        source_type,
                        batch_name,
                    ),
                )
            loaded += 1
        except Exception as exc:
            failed += 1
            cur.execute(
                """
                INSERT INTO load_errors (
                    source_table, raw_record_id, error_type, error_message, raw_data, is_resolved
                ) VALUES (?, NULL, 'parse_error', ?, ?, 0)
                """,
                (f"{entity}_raw", str(exc), json.dumps(row, ensure_ascii=False)),
            )

    cur.execute(
        """
        UPDATE source_import_runs
        SET records_total = ?, records_loaded = ?, records_failed = ?, status = ?, finished_at = ?
        WHERE id = ?
        """,
        (
            len(rows),
            loaded,
            failed,
            "completed" if failed == 0 else "completed_with_errors",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            run_id,
        ),
    )
    conn.commit()
    return loaded, failed


def map_hotel_row(row, source_type):
    if source_type == "csv":
        return {
            "raw_name": row["hotel_name"],
            "raw_country": row["country"],
            "raw_resort": row["resort"],
            "raw_stars": row["stars"],
            "raw_meal": row["meal"],
        }
    if source_type == "json":
        return {
            "raw_name": row["title"],
            "raw_country": row["destination_country"],
            "raw_resort": row["destination_resort"],
            "raw_stars": row["stars"],
            "raw_meal": row["board"],
        }
    return {
        "raw_name": row["HotelName"],
        "raw_country": row["CountryName"],
        "raw_resort": row["ResortName"],
        "raw_stars": row["Stars"],
        "raw_meal": row["MealPlan"],
    }


def map_tour_row(row, source_type):
    if source_type == "csv":
        return {
            "raw_hotel_name": row["hotel_name"],
            "raw_country": row["country"],
            "raw_resort": row["resort"],
            "raw_price": row["price"],
            "raw_departure": row["departure_date"],
            "raw_duration": row["nights"],
            "raw_meal": row["meal"],
            "raw_adults": row["adults"],
            "raw_children": row["children"],
        }
    if source_type == "json":
        return {
            "raw_hotel_name": row["hotel"],
            "raw_country": row["countryName"],
            "raw_resort": row["resortName"],
            "raw_price": row["amount"],
            "raw_departure": row["startDate"],
            "raw_duration": row["duration"],
            "raw_meal": row["mealCode"],
            "raw_adults": row["adultsCount"],
            "raw_children": row["childrenCount"],
        }
    return {
        "raw_hotel_name": row["HotelName"],
        "raw_country": row["CountryName"],
        "raw_resort": row["ResortName"],
        "raw_price": row["PriceValue"],
        "raw_departure": row["DepartureDate"],
        "raw_duration": row["Nights"],
        "raw_meal": row["MealPlan"],
        "raw_adults": row["Adults"],
        "raw_children": row["Children"],
    }


def create_database():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    os.makedirs(SOURCE_DIR, exist_ok=True)
    safe_reset_path(DB_PATH)
    safe_reset_path(SOURCE_DIR)
    os.makedirs(SOURCE_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    with open(SCHEMA_PATH, "r", encoding="utf-8") as fh:
        cur.executescript(fh.read())
    log("[OK] Схема БД создана")

    country_ids = {}
    for name, iso, region, popular in COUNTRIES:
        cur.execute(
            "INSERT INTO countries (name, iso_code, region, is_popular) VALUES (?, ?, ?, ?)",
            (name, iso, region, popular),
        )
        country_ids[name] = cur.lastrowid
    log(f"[OK] Страны: {len(country_ids)}")

    resort_ids = {}
    for country, resort_list in RESORTS_BY_COUNTRY.items():
        for resort in resort_list:
            cur.execute(
                "INSERT INTO resorts (country_id, name, is_coastal, avg_temp_summer) VALUES (?, ?, ?, ?)",
                (country_ids[country], resort, 1 if random.random() < 0.7 else 0, round(random.uniform(24, 35), 1)),
            )
            resort_ids[(country, resort)] = cur.lastrowid
    log(f"[OK] Курорты: {len(resort_ids)}")

    operator_ids = {}
    operator_by_short = {}
    for name, short, active, website in OPERATORS:
        cur.execute(
            "INSERT INTO tour_operators (name, short_name, is_active, website) VALUES (?, ?, ?, ?)",
            (name, short, active, website),
        )
        operator_ids[name] = cur.lastrowid
        operator_by_short[short] = operator_ids[name]
    log(f"[OK] Операторы: {len(operator_ids)}")

    mapping_count = 0
    for op_name, raw_names in OPERATOR_RAW_NAMES.items():
        oid = operator_ids[op_name]
        for i, raw_name in enumerate(raw_names):
            cur.execute(
                "INSERT INTO operator_mapping (operator_id, raw_name, source, confidence) VALUES (?, ?, ?, ?)",
                (oid, raw_name, ["csv", "json", "xlsx", "scraping", "manual"][i % 5], 1.0 if i == 0 else round(random.uniform(0.8, 0.99), 2)),
            )
            mapping_count += 1
    log(f"[OK] Маппинг операторов: {mapping_count}")

    geo_alias_count = 0
    for canonical, aliases in COUNTRY_VARIANTS.items():
        for alias in aliases:
            cur.execute(
                """
                INSERT INTO geo_aliases (entity_type, canonical_name, alias_name, country_id, source, confidence)
                VALUES ('country', ?, ?, ?, 'seed', 1.0)
                """,
                (canonical, alias, country_ids[canonical]),
            )
            geo_alias_count += 1
    for canonical, aliases in RESORT_VARIANTS.items():
        country_name = next(country for country, resorts in RESORTS_BY_COUNTRY.items() if canonical in resorts)
        for alias in aliases:
            cur.execute(
                """
                INSERT INTO geo_aliases (entity_type, canonical_name, alias_name, country_id, resort_id, source, confidence)
                VALUES ('resort', ?, ?, ?, ?, 'seed', 1.0)
                """,
                (canonical, alias, country_ids[country_name], resort_ids[(country_name, canonical)]),
            )
            geo_alias_count += 1
    log(f"[OK] Алиасы географии: {geo_alias_count}")

    hotels = []
    hotel_alias_count = 0
    for country, resort_list in RESORTS_BY_COUNTRY.items():
        for resort in resort_list:
            rid = resort_ids[(country, resort)]
            for _ in range(random.randint(8, 18)):
                name = f"{random.choice(HOTEL_NAME_PARTS['prefix'])} {random.choice(HOTEL_NAME_PARTS['name'])}"
                stars = random.choices([2, 3, 4, 5], weights=[5, 20, 50, 25])[0]
                meal = random.choice(MEAL_TYPES)
                lat = round(random.uniform(25, 55), 4)
                lon = round(random.uniform(20, 140), 4)
                cur.execute(
                    """
                    INSERT INTO hotels (resort_id, name, star_rating, meal_plan, latitude, longitude)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (rid, name, stars, meal, lat, lon),
                )
                hotel_id = cur.lastrowid
                cur.execute(
                    """
                    INSERT INTO hotel_aliases (hotel_id, alias_name, source, confidence)
                    VALUES (?, ?, 'canonical', 1.0)
                    """,
                    (hotel_id, name),
                )
                hotel_alias_count += 1
                hotels.append(
                    {
                        "id": hotel_id,
                        "name": name,
                        "country": country,
                        "country_id": country_ids[country],
                        "resort": resort,
                        "resort_id": rid,
                        "stars": stars,
                        "meal": meal,
                    }
                )
    log(f"[OK] Отели: {len(hotels)}")
    log(f"[OK] Алиасы отелей: {hotel_alias_count}")

    today = datetime.now()
    season_end = today + timedelta(days=180)
    tours = []
    for hotel in hotels:
        for _ in range(random.randint(2, 4)):
            operator_name, operator_short, _, _ = random.choice(OPERATORS)
            departure = random_date(today, season_end)
            nights = random.choice([5, 6, 7, 8, 9, 10, 11, 12, 14])
            meal = random.choice(MEAL_TYPES)
            price = generate_tour_price(hotel["country"], hotel["stars"], meal, nights, operator_short)
            discount_pct = OPERATOR_DISCOUNTS[operator_short] * 100
            original_price = round(price / (1 - discount_pct / 100), 0) if discount_pct else price
            cur.execute(
                """
                INSERT INTO tours (
                    operator_id, hotel_id, country_id, resort_id, departure_date, duration_nights,
                    meal_type, actual_price, original_price, discount_pct, adults, children, room_type,
                    is_available, source_url, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    operator_ids[operator_name],
                    hotel["id"],
                    hotel["country_id"],
                    hotel["resort_id"],
                    departure.strftime("%Y-%m-%d"),
                    nights,
                    meal,
                    price,
                    original_price,
                    discount_pct,
                    2,
                    random.choices([0, 1, 2], weights=[55, 30, 15])[0],
                    random.choice(ROOM_TYPES),
                    1 if random.random() < 0.92 else 0,
                    f"https://example.test/{operator_short.lower()}/tour/{hotel['id']}",
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            tour_id = cur.lastrowid
            tours.append(
                {
                    "id": tour_id,
                    "operator_id": operator_ids[operator_name],
                    "operator_name": operator_name,
                    "operator_short": operator_short,
                    "hotel_id": hotel["id"],
                    "hotel_name": hotel["name"],
                    "country": hotel["country"],
                    "resort": hotel["resort"],
                    "departure": departure,
                    "nights": nights,
                    "meal": meal,
                    "price": price,
                    "children": cur.execute("SELECT children FROM tours WHERE id = ?", (tour_id,)).fetchone()[0],
                }
            )
    log(f"[OK] Туры: {len(tours)}")

    price_history_count = 0
    for tour in random.sample(tours, min(900, len(tours))):
        for _ in range(random.randint(3, 7)):
            recorded_at = today - timedelta(days=random.randint(1, 60))
            hist_price = round(tour["price"] * random.uniform(0.87, 1.13), 0)
            cur.execute(
                "INSERT INTO price_history (tour_id, price, recorded_at) VALUES (?, ?, ?)",
                (tour["id"], hist_price, recorded_at.strftime("%Y-%m-%d %H:%M:%S")),
            )
            price_history_count += 1
    log(f"[OK] История цен: {price_history_count}")

    users = []
    for idx in range(500):
        cur.execute(
            "INSERT INTO users (username, email, city, age_group, registered_at) VALUES (?, ?, ?, ?, ?)",
            (
                f"user_{idx + 1}",
                f"user_{idx + 1}@mail.ru",
                random.choice(CITIES),
                random.choice(AGE_GROUPS),
                random_date(datetime(2023, 1, 1), datetime(2025, 12, 31)).strftime("%Y-%m-%d"),
            ),
        )
        users.append(cur.lastrowid)
    log(f"[OK] Пользователи: {len(users)}")

    source_stats = create_operator_source_files(conn, tours, hotels, operator_ids)
    log(f"[OK] Источники операторов: {source_stats['files']} файлов")
    log(f"[OK] Загружено сырых отелей: {source_stats['raw_hotels']}")
    log(f"[OK] Загружено сырых туров: {source_stats['raw_tours']}")

    create_user_funnel(cur, users, tours, country_ids, resort_ids)
    seed_system_logs(cur)

    conn.commit()
    conn.close()
    log(f"\n[READY] База данных создана: {DB_PATH}")
    log(f"        Размер: {os.path.getsize(DB_PATH) / 1024:.0f} KB")


def create_operator_source_files(conn, tours, hotels, operator_ids):
    source_types = ["csv", "json", "xlsx"]
    total_files = 0
    raw_hotels = 0
    raw_tours = 0
    batch_name = "batch_001"

    for idx, (operator_name, _short, _active, _website) in enumerate(OPERATORS):
        source_type = source_types[idx % len(source_types)]
        operator_id = operator_ids[operator_name]

        hotel_rows = []
        for hotel in random.sample(hotels, min(180, len(hotels))):
            hotel_rows.append(build_source_hotel_row(hotel, source_type))

        tour_rows = []
        operator_tours = [tour for tour in tours if tour["operator_id"] == operator_id]
        for tour in random.sample(operator_tours, min(260, len(operator_tours))):
            tour_rows.append(build_source_tour_row(tour, source_type))
        for _ in range(25):
            tour_rows.append(build_invalid_tour_row(source_type))

        hotel_path = os.path.join(SOURCE_DIR, f"{operator_name.replace('&', 'and').replace(' ', '_').lower()}_hotels.{source_type}")
        tour_path = os.path.join(SOURCE_DIR, f"{operator_name.replace('&', 'and').replace(' ', '_').lower()}_tours.{source_type}")

        save_source_file(hotel_path, source_type, "hotels", hotel_rows)
        save_source_file(tour_path, source_type, "tours", tour_rows)
        total_files += 2

        loaded_hotels, _ = import_source_rows(conn, operator_id, operator_name, source_type, "hotels", hotel_path, batch_name)
        loaded_tours, _ = import_source_rows(conn, operator_id, operator_name, source_type, "tours", tour_path, batch_name)
        raw_hotels += loaded_hotels
        raw_tours += loaded_tours

    return {"files": total_files, "raw_hotels": raw_hotels, "raw_tours": raw_tours}


def build_source_hotel_row(hotel, source_type):
    payload = {
        "name": create_variant_hotel_name(hotel["name"]),
        "country": choose_variant(COUNTRY_VARIANTS, hotel["country"]),
        "resort": choose_variant(RESORT_VARIANTS, hotel["resort"]) if hotel["resort"] in RESORT_VARIANTS else hotel["resort"],
        "stars": random.choice([str(hotel["stars"]), f"{hotel['stars']}.0", "пять" if hotel["stars"] == 5 else str(hotel["stars"])]),
        "meal": random.choice(MEAL_VARIANTS[hotel["meal"]]),
    }
    if random.random() < 0.08:
        payload["country"] = random.choice(["", "N/A", "Таеланд", "ОАЕ"])
    if random.random() < 0.1:
        payload["meal"] = random.choice(["", "???", "All"])
    if source_type == "csv":
        return {
            "hotel_name": payload["name"],
            "country": payload["country"],
            "resort": payload["resort"],
            "stars": payload["stars"],
            "meal": payload["meal"],
        }
    if source_type == "json":
        return {
            "title": payload["name"],
            "destination_country": payload["country"],
            "destination_resort": payload["resort"],
            "stars": payload["stars"],
            "board": payload["meal"],
        }
    return {
        "HotelName": payload["name"],
        "CountryName": payload["country"],
        "ResortName": payload["resort"],
        "Stars": payload["stars"],
        "MealPlan": payload["meal"],
    }


def build_source_tour_row(tour, source_type):
    payload = {
        "hotel": create_variant_hotel_name(tour["hotel_name"]),
        "country": choose_variant(COUNTRY_VARIANTS, tour["country"]),
        "resort": choose_variant(RESORT_VARIANTS, tour["resort"]) if tour["resort"] in RESORT_VARIANTS else tour["resort"],
        "price": str(int(tour["price"])),
        "departure": tour["departure"].strftime("%Y-%m-%d"),
        "nights": str(tour["nights"]),
        "meal": random.choice(MEAL_VARIANTS[tour["meal"]]),
        "adults": "2",
        "children": str(tour["children"]),
    }
    if source_type == "csv":
        return {
            "hotel_name": payload["hotel"],
            "country": payload["country"],
            "resort": payload["resort"],
            "price": payload["price"],
            "departure_date": payload["departure"],
            "nights": payload["nights"],
            "meal": payload["meal"],
            "adults": payload["adults"],
            "children": payload["children"],
        }
    if source_type == "json":
        return {
            "hotel": payload["hotel"],
            "countryName": payload["country"],
            "resortName": payload["resort"],
            "amount": payload["price"],
            "startDate": payload["departure"],
            "duration": payload["nights"],
            "mealCode": payload["meal"],
            "adultsCount": payload["adults"],
            "childrenCount": payload["children"],
        }
    return {
        "HotelName": payload["hotel"],
        "CountryName": payload["country"],
        "ResortName": payload["resort"],
        "PriceValue": payload["price"],
        "DepartureDate": payload["departure"],
        "Nights": payload["nights"],
        "MealPlan": payload["meal"],
        "Adults": payload["adults"],
        "Children": payload["children"],
    }


def build_invalid_tour_row(source_type):
    error_kind = random.choice(["bad_date", "zero_price", "negative_price", "missing_field", "bad_duration"])
    payload = {
        "hotel": random.choice(["", "UNKNOWN HOTEL", "HTL ???"]),
        "country": random.choice(["", "Турций", "N/A"]),
        "resort": random.choice(["", "не указан", "N/A"]),
        "price": random.choice(["0", "-55000", "not_a_number"]),
        "departure": random.choice(["2024-13-01", "", "не указана", "NaN"]),
        "nights": random.choice(["0", "-3", "неделя", ""]),
        "meal": random.choice(["", "FB+", "Всё!", "??"]),
        "adults": random.choice(["0", "2", "8"]),
        "children": random.choice(["0", "-1", "5"]),
    }
    if error_kind == "zero_price":
        payload["price"] = "0"
    elif error_kind == "negative_price":
        payload["price"] = "-55000"
    elif error_kind == "missing_field":
        payload["hotel"] = ""
    elif error_kind == "bad_duration":
        payload["nights"] = "неделя"

    if source_type == "csv":
        return {
            "hotel_name": payload["hotel"],
            "country": payload["country"],
            "resort": payload["resort"],
            "price": payload["price"],
            "departure_date": payload["departure"],
            "nights": payload["nights"],
            "meal": payload["meal"],
            "adults": payload["adults"],
            "children": payload["children"],
        }
    if source_type == "json":
        return {
            "hotel": payload["hotel"],
            "countryName": payload["country"],
            "resortName": payload["resort"],
            "amount": payload["price"],
            "startDate": payload["departure"],
            "duration": payload["nights"],
            "mealCode": payload["meal"],
            "adultsCount": payload["adults"],
            "childrenCount": payload["children"],
        }
    return {
        "HotelName": payload["hotel"],
        "CountryName": payload["country"],
        "ResortName": payload["resort"],
        "PriceValue": payload["price"],
        "DepartureDate": payload["departure"],
        "Nights": payload["nights"],
        "MealPlan": payload["meal"],
        "Adults": payload["adults"],
        "Children": payload["children"],
    }


def save_source_file(path, source_type, entity, rows):
    if source_type == "csv":
        fieldnames = list(rows[0].keys())
        write_csv(path, rows, fieldnames)
        return
    if source_type == "json":
        write_json(path, rows)
        return
    headers = list(rows[0].keys())
    write_xlsx(path, entity, rows, headers)


def create_user_funnel(cur, users, tours, country_ids, resort_ids):
    search_count = 0
    view_count = 0
    click_count = 0
    booking_count = 0
    sessions = []
    tour_ids = [tour["id"] for tour in tours]

    for idx in range(2000):
        session_id = f"sess_{idx + 1}"
        user_id = random.choice(users) if random.random() < 0.82 else None
        country = random.choice(list(country_ids.keys()))
        resort_options = RESORTS_BY_COUNTRY[country]
        resort = random.choice(resort_options)
        date_from = random_date(datetime.now(), datetime.now() + timedelta(days=120))
        date_to = date_from + timedelta(days=random.choice([5, 7, 10, 12, 14]))
        cur.execute(
            """
            INSERT INTO search_logs (
                user_id, session_id, country_id, resort_id, date_from, date_to,
                adults, children, max_price, min_duration, results_count, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                session_id,
                country_ids[country],
                resort_ids[(country, resort)],
                date_from.strftime("%Y-%m-%d"),
                date_to.strftime("%Y-%m-%d"),
                random.choice([1, 2, 2, 3, 4]),
                random.choice([0, 0, 1, 2]),
                random.choice([50000, 80000, 100000, 120000, 150000, 200000, 300000]),
                random.choice([5, 6, 7, 10]),
                random.randint(5, 150),
                random_date(datetime(2025, 1, 1), datetime(2025, 12, 31)).strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        search_count += 1
        sessions.append({"session_id": session_id, "user_id": user_id})

    for session in sessions:
        viewed_tours = []
        if random.random() < 0.76:
            for _ in range(random.randint(1, 4)):
                tour_id = random.choice(tour_ids)
                viewed_tours.append(tour_id)
                cur.execute(
                    "INSERT INTO tour_views (user_id, session_id, tour_id, view_duration, created_at) VALUES (?, ?, ?, ?, ?)",
                    (
                        session["user_id"],
                        session["session_id"],
                        tour_id,
                        random.randint(5, 320),
                        random_date(datetime(2025, 1, 1), datetime(2025, 12, 31)).strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                view_count += 1

        clicked_tour = None
        if viewed_tours and random.random() < 0.47:
            for _ in range(random.randint(1, 2)):
                clicked_tour = random.choice(viewed_tours)
                cur.execute(
                    "INSERT INTO click_logs (user_id, session_id, tour_id, click_type, created_at) VALUES (?, ?, ?, ?, ?)",
                    (
                        session["user_id"],
                        session["session_id"],
                        clicked_tour,
                        random.choice(["details", "book", "compare"]),
                        random_date(datetime(2025, 1, 1), datetime(2025, 12, 31)).strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                click_count += 1

        if clicked_tour and random.random() < 0.22:
            cur.execute("SELECT actual_price FROM tours WHERE id = ?", (clicked_tour,))
            total_price = cur.fetchone()[0]
            cur.execute(
                """
                INSERT INTO bookings (
                    user_id, session_id, tour_id, booking_date, total_price, status,
                    adults, children, contact_email, contact_phone
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    session["user_id"] or random.choice(users),
                    session["session_id"],
                    clicked_tour,
                    random_date(datetime(2025, 1, 1), datetime(2025, 12, 31)).strftime("%Y-%m-%d"),
                    total_price,
                    random.choices(BOOKING_STATUSES, weights=[15, 45, 20, 20])[0],
                    2,
                    random.choice([0, 0, 1, 2]),
                    f"booking_{booking_count + 1}@mail.ru",
                    f"+7{random.randint(9000000000, 9999999999)}",
                ),
            )
            booking_count += 1

    log(f"[OK] Логи поиска: {search_count}")
    log(f"[OK] Просмотры: {view_count}")
    log(f"[OK] Клики: {click_count}")
    log(f"[OK] Бронирования: {booking_count}")


def seed_system_logs(cur):
    log_rows = [
        ("INFO", "loader", "Запущена загрузка данных операторов", "CSV / JSON / XLSX"),
        ("INFO", "normalization", "Подготовлены справочники алиасов", "countries,resorts,hotels"),
        ("WARN", "loader", "Часть raw-туров содержит ошибки", "Ожидается разбор на этапе нормализации"),
    ]
    for level, module, message, details in log_rows:
        cur.execute(
            "INSERT INTO system_logs (log_level, module, message, details) VALUES (?, ?, ?, ?)",
            (level, module, message, details),
        )


if __name__ == "__main__":
    random.seed(42)
    create_database()
